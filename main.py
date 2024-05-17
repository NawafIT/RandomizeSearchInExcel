from random import choice
import openpyxl


wb = openpyxl.load_workbook("Responses.xlsx")
sh1 = wb.active

# index starts from 1
column = 4
column_data = [cell for row in sh1.iter_cols(min_col=column, max_col=column, min_row=1, max_row=sh1.max_row, values_only=True) for cell in row]

column_data = [cell for cell in column_data if cell is not None]

print(choice(column_data))
